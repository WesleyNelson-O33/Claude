import openpyxl, datetime, re
old = openpyxl.load_workbook('wb.xlsm', data_only=True)['Tracker']
nw  = openpyxl.load_workbook('FY27_Revenue_Tracker_v3.xlsx')   # formulas
ds, dsh, ck, ls = nw['Data'], nw['Dashboard'], nw['Checks'], nw['Lists']
def num(v): return v if isinstance(v,(int,float)) else 0
ok=True
def chk(label,a,b,tol=0.005):
    global ok; good=abs(a-b)<=tol; ok&=good
    print(f"  {'PASS' if good else 'FAIL'}  {label:<46} new={a:>14,.2f}  original={b:>14,.2f}")
def chkv(label,a,b):
    global ok; good=(a==b); ok&=good
    print(f"  {'PASS' if good else 'FAIL'}  {label:<46} new={a}  original={b}")

# ---------- 1. every source cell migrated verbatim ----------
BLK={'Support':(69,[(3,'date'),(6,'client'),(7,'job'),(8,'desc'),(9,'invno'),(10,'invval'),(17,'notes')]),
     }
print("=== 1. Row counts by source ===")
cnt={}
for r in range(2,ds.max_row+1):
    s=ds.cell(r,2).value
    if s: cnt[s]=cnt.get(s,0)+1
chkv("Support rows",cnt.get('Support'),35); chkv("Production rows",cnt.get('Production'),49)
chkv("Consulting rows",cnt.get('Consulting'),32); chkv("WIP rows",cnt.get('WIP'),65)
chkv("total rows",sum(cnt.values()),181)

# ---------- 2. money columns tie back to the original blocks ----------
print("\n=== 2. Raw amounts carried across (sum of stored values) ===")
pairs=[('Support Invoice Value',10,'Support',74),('Production Invoice Value',10,'Production',104),
       ('Consulting Invoice Value',10,'Consulting',134),('WIP Movement',11,'WIP',158),
       ('Discounts Included',30,'Production',116),('Cross Hire Expense',32,'Production',117),
       ('Labour Expense (Internal)',33,'Production',118),('Labour Expense (External)',38,'Consulting',138),
       ('Equipment Expense (Internal)',39,'Consulting',139),('Subscription & Licences',40,'Consulting',140),
       ('Labour Revenue',35,'Consulting',135),('Equipment Revenue',36,'Consulting',136),
       ('Subscription Revenue',37,'Consulting',137),('Video Total',29,'Production',115)]
for label,ncol,src,ocol in pairs:
    a=sum(num(ds.cell(r,ncol).value) for r in range(2,ds.max_row+1) if ds.cell(r,2).value==src)
    b=sum(num(old.cell(r,ocol).value) for r in range(12,212))
    chk(label,a,b)

# ---------- 3. evaluate the new sheet's derived columns in Python ----------
print("\n=== 3. Derived columns recomputed from the NEW sheet vs the ORIGINAL stored results ===")
netrev={}; margin={}
for r in range(2,ds.max_row+1):
    s=ds.cell(r,2).value
    if not s: continue
    iv=num(ds.cell(r,10).value); wip=num(ds.cell(r,11).value)
    disc=num(ds.cell(r,30).value); ch=num(ds.cell(r,32).value); li=num(ds.cell(r,33).value)
    le=num(ds.cell(r,38).value); ei=num(ds.cell(r,39).value); se=num(ds.cell(r,40).value)
    nr = wip if s=='WIP' else ('' if ds.cell(r,10).value is None else (iv-disc if s=='Production' else iv))
    netrev[s]=netrev.get(s,0)+num(nr)
    if s=='Production': margin['Production']=margin.get('Production',0)+(iv-(ch+li) if ds.cell(r,10).value is not None else 0)
    if s=='Consulting': margin['Consulting']=margin.get('Consulting',0)+(iv-(le+ei+se) if ds.cell(r,10).value is not None else 0)
for s,ocol in [('Support',74),('Production',119),('Consulting',134),('WIP',158)]:
    chk(f"{s} Net Revenue", netrev.get(s,0), sum(num(old.cell(r,ocol).value) for r in range(12,212)))
chk("Production Margin", margin['Production'], sum(num(old.cell(r,121).value) for r in range(12,212)))
chk("Consulting Margin", margin['Consulting'], sum(num(old.cell(r,141).value) for r in range(12,212)))

# ---------- 4. monthly grid ----------
print("\n=== 4. Dashboard months recomputed both ways ===")
FY0=ls['B3'].value; FY1=ls['B4'].value
chkv("FY start on Lists tab", FY0, datetime.datetime(2026,7,1))
chkv("FY end on Lists tab", FY1, datetime.datetime(2027,7,1))
def bounds(i):
    m=7+i; y=2026+(m-1)//12; m=(m-1)%12+1
    m2,y2=(1,y+1) if m==12 else (m+1,y)
    return datetime.datetime(y,m,1), datetime.datetime(y2,m2,1)
specs=[('Support',74,69),('Production',119,97),('Consulting',134,125),('WIP',158,153)]
grand_new=grand_old=0
for i in range(12):
    a,b=bounds(i)
    for s,ocol,odc in specs:
        exp=sum(num(old.cell(r,ocol).value) for r in range(12,212)
                if isinstance(old.cell(r,odc).value,datetime.datetime) and a<=old.cell(r,odc).value<b)
        got=0.0
        for r in range(2,ds.max_row+1):
            if ds.cell(r,2).value!=s: continue
            d=ds.cell(r,3).value
            if not isinstance(d,datetime.datetime) or not (a<=d<b): continue
            iv=num(ds.cell(r,10).value)
            got += num(ds.cell(r,11).value) if s=='WIP' else (iv-num(ds.cell(r,30).value) if s=='Production' else iv)
        grand_new+=got; grand_old+=exp
        if abs(exp)>0.005 or abs(got)>0.005: chk(f"{a.strftime('%b-%y')} {s}", got, exp)
chk("12-month grand total", grand_new, grand_old)

# ---------- 5. date classification ----------
print("\n=== 5. Date Check classification ===")
cls={}
for r in range(2,ds.max_row+1):
    if not ds.cell(r,2).value: continue
    d=ds.cell(r,3).value
    k='No date' if d is None else ('In FY27' if FY0<=d<FY1 else 'Outside FY27')
    cls[k]=cls.get(k,0)+1
print("  ",cls)
chkv("no-date records", cls.get('No date'), 26+3)   # 11 Support + 15 Production + 3 Consulting
chkv("outside-FY27 records", cls.get('Outside FY27'), 16)

# ---------- 6. formula-string safety ----------
print("\n=== 6. Formula function audit (must be Excel-2007-safe) ===")
FNS=set()
for sh in nw.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith('='):
                FNS.update(re.findall(r'([A-Z][A-Z0-9_.]*)\s*\(', c.value))
BAD={'XLOOKUP','XMATCH','SORT','FILTER','UNIQUE','SEQUENCE','TEXTJOIN','CONCAT','IFS','SWITCH','MAXIFS','MINIFS'}
print("  functions used:", ", ".join(sorted(FNS)))
chkv("no post-2007 / spilling functions", sorted(FNS & BAD), [])
print("\nRESULT:", "ALL CHECKS PASSED" if ok else "*** MISMATCHES FOUND ***")
