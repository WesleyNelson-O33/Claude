import json, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as L

D = json.load(open('extract.json'))
def dec(o):
    if isinstance(o, dict) and '__d' in o: return datetime.datetime.fromisoformat(o['__d'])
    return o
recs = [{k: dec(v) for k, v in r.items()} for r in D['records']]

NAVY   = '3E5066'; LIGHT = '5B708A'; GREY = 'EDEDED'
BAND   = 'F5F7F9'; YELLOW= 'FFF3C4'; RULE = 'B7B7B7'
F      = 'Arial'
MONEY  = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PCT    = '0.0%'; DATE = 'dd-mmm-yy'
thin   = Side(style='thin', color=RULE)
box    = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ==========================================================  LISTS
ls = wb.active; ls.title = 'Lists'
ls['A1'] = 'Control values — edit here and every dropdown and check updates'
ls['A1'].font = Font(F, 11, bold=True, color=NAVY)
ls['A3'] = 'Financial year start';       ls['B3'] = datetime.datetime(2026, 7, 1)
ls['A4'] = 'Financial year end (excl.)'; ls['B4'] = datetime.datetime(2027, 7, 1)
for c in ('B3', 'B4'):
    ls[c].number_format = DATE; ls[c].font = Font(F, 10, color='0000FF'); ls[c].fill = PatternFill('solid', fgColor=YELLOW)

cols = {'D': ('Source', ['Support', 'Production', 'Consulting', 'WIP']),
        'F': ('Department', ['Consulting', 'Integration', 'Onsite', 'Other', 'Production', 'Support', 'Video']),
        'H': ('Status', ['In Progress', 'Completed', 'Cancelled', 'Closed']),
        'J': ('Yes / No', ['Y', 'N'])}
for col, (title, items) in cols.items():
    ls[f'{col}3'] = title
    ls[f'{col}3'].font = Font(F, 10, bold=True, color='FFFFFF')
    ls[f'{col}3'].fill = PatternFill('solid', fgColor=NAVY)
    ls[f'{col}3'].alignment = Alignment(horizontal='center')
    for i, x in enumerate(items):
        c = ls[f'{col}{4+i}']; c.value = x; c.font = Font(F, 10); c.border = box
for col in ('A', 'D', 'F', 'H', 'J'): ls.column_dimensions[col].width = 24
ls.column_dimensions['B'].width = 14
for r in ls['A1:K30']:
    for c in r:
        if c.font.name != F: c.font = Font(F, 10)

# ==========================================================  DATA
ds = wb.create_sheet('Data')
HDRS = [
 ('No.', 6), ('Source', 12), ('Date', 11), ('Department', 13), ('Status', 12),
 ('Client', 22), ('Job Number', 13), ('Description', 40), ('Invoice Number', 15),
 ('Invoice Value', 14), ('WIP Movement', 14), ('Net Revenue', 14),
 ('Margin', 13), ('Margin %', 10), ('Date Check', 13), ('Duplicate Check', 19), ('Notes', 34),
 ('Client Email', 26), ('Event Date', 11), ('Zoho Number', 13), ('Current Number', 14),
 ('Closed', 9), ('Invoice Posted to Xero', 13), ('Video Filming', 11), ('Video Editing', 11),
 ('Project Management', 12), ('Video Project Management', 13), ('Production Labour Hours', 12),
 ('Video Total', 13), ('Discounts Included', 13), ('Discounts as % of Net', 12),
 ('Cross Hire Expense', 14), ('Labour Expense (Internal)', 14),
 ('Qwilr Link', 34), ('Labour Revenue', 14), ('Equipment Revenue', 14), ('Subscription Revenue', 14),
 ('Labour Expense (External)', 14), ('Equipment Expense (Internal)', 14), ('Subscription & Licences Expense', 14),
]
for i, (h, w) in enumerate(HDRS, 1):
    c = ds.cell(1, i, h)
    c.font = Font(F, 10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=(LIGHT if 18 <= i <= 33 else NAVY))
    c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    c.border = box
    ds.column_dimensions[L(i)].width = w
ds.row_dimensions[1].height = 46

KEYS = ['', 'source', 'date', 'dept', 'status', 'client', 'job', 'desc', 'invno',
        'invval', 'wip', '', '', '', '', '', 'notes',
        'email', 'eventdate', 'zoho', 'curnum', 'closed', 'posted', 'vfilm', 'vedit',
        'pm', 'vpm', 'hours', 'vtotal', 'disc', '', 'crosshire', 'labint',
        'qwilr', 'labrev', 'eqprev', 'subrev', 'labext', 'eqpint', 'subexp']

order = {'Support': 0, 'Production': 1, 'Consulting': 2, 'WIP': 3}
recs.sort(key=lambda r: (order[r['source']], r['_src_row']))
N = len(recs)
for i, rec in enumerate(recs):
    r = i + 2
    for ci, key in enumerate(KEYS, 1):
        if key and rec.get(key) is not None:
            ds.cell(r, ci, rec[key])
    ds.cell(r, 1,  f'=IF(COUNTA($B{r}:$K{r})=0,"",ROW()-1)')
    ds.cell(r, 12, f'=IF($B{r}="WIP",$K{r},IF($J{r}="","",IF($B{r}="Production",$J{r}-N($AD{r}),$J{r})))')
    ds.cell(r, 13, f'=IF($B{r}="Production",IF($J{r}="","",$J{r}-(N($AF{r})+N($AG{r}))),'
                   f'IF($B{r}="Consulting",IF($J{r}="","",$J{r}-(N($AL{r})+N($AM{r})+N($AN{r}))),""))')
    ds.cell(r, 14, f'=IFERROR($M{r}/$J{r},"")')
    ds.cell(r, 15, f'=IF(COUNTA($B{r}:$K{r})=0,"",IF($C{r}="","No date",'
                   f'IF(AND($C{r}>=Lists!$B$3,$C{r}<Lists!$B$4),"In FY27","Outside FY27")))')
    ds.cell(r, 16, f'=IF($B{r}<>"WIP","",IF($G{r}="","",'
                   f'IF(COUNTIFS($B$2:$B$20000,"<>WIP",$G$2:$G$20000,$G{r})>0,"Job already invoiced","")))')
    ds.cell(r, 31, f'=IF($B{r}<>"Production","",IFERROR($AD{r}/$L{r},""))')

LAST = N + 1
MONEY_COLS = [10, 11, 12, 13, 29, 30, 32, 33, 35, 36, 37, 38, 39, 40]
PCT_COLS   = [14, 31]
DATE_COLS  = [3, 19]
CTR_COLS   = [1, 2, 3, 7, 9, 15, 16, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
for r in range(2, LAST + 1):
    ds.row_dimensions[r].height = 15
    for ci in range(1, len(HDRS) + 1):
        c = ds.cell(r, ci)
        c.font = Font(F, 10); c.border = box
        c.alignment = Alignment(horizontal=('center' if ci in CTR_COLS else 'left'), vertical='center')
        if ci in MONEY_COLS: c.number_format = MONEY
        elif ci in PCT_COLS: c.number_format = PCT
        elif ci in DATE_COLS: c.number_format = DATE
        if ci in (1, 12, 13, 14, 15, 16, 31): c.fill = PatternFill('solid', fgColor=GREY)

tbl = Table(displayName='tblRevenue', ref=f'A1:{L(len(HDRS))}{LAST}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False, showColumnStripes=False)
ds.add_table(tbl)

for rng, formula in ((f'B2:B{LAST}', '=Lists!$D$4:$D$7'), (f'D2:D{LAST}', '=Lists!$F$4:$F$10'),
                     (f'E2:E{LAST}', '=Lists!$H$4:$H$7'), (f'V2:W{LAST}', '=Lists!$J$4:$J$5')):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = 'Pick a value from the list on the Lists tab.'
    ds.add_data_validation(dv); dv.add(rng)

ds.column_dimensions.group('R', 'AG', outline_level=1, hidden=False)
ds.column_dimensions.group('AH', 'AN', outline_level=1, hidden=False)
ds.sheet_properties.outlinePr.summaryRight = False
ds.freeze_panes = 'D2'
ds.sheet_view.showGridLines = False

# ==========================================================  DASHBOARD
dsh = wb.create_sheet('Dashboard')
dsh.sheet_view.showGridLines = False
def title(ws, cell, text, size=13):
    ws[cell] = text; ws[cell].font = Font(F, size, bold=True, color=NAVY)
def hdrow(ws, row, labels, start=1, fill=NAVY):
    for i, t in enumerate(labels):
        c = ws.cell(row, start + i, t)
        c.font = Font(F, 10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=fill)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = box
    ws.row_dimensions[row].height = 30

title(dsh, 'A1', 'FY 2026-27 — Monthly revenue by department')
hdrow(dsh, 2, ['Month', 'Support', 'Production', 'Consulting', 'WIP',
               'Total Revenue', 'P&L Total', 'Variance'])
for i in range(12):
    r = 3 + i
    m = 7 + i; y = 2026 + (m - 1) // 12; m = (m - 1) % 12 + 1
    dsh.cell(r, 1, datetime.datetime(y, m, 1)).number_format = 'mmm-yy'
    for ci in range(2, 6):
        dsh.cell(r, ci, f'=SUMIFS(Data!$L$2:$L$200000,Data!$B$2:$B$20000,{L(ci)}$2,'
                        f'Data!$C$2:$C$20000,">="&$A{r},Data!$C$2:$C$20000,"<"&EDATE($A{r},1))')
    dsh.cell(r, 6, f'=SUM($B{r}:$E{r})')
    dsh.cell(r, 8, f'=IF($G{r}="","",IF(ABS($F{r}-$G{r})<0.99,"Reconciled",$F{r}-$G{r}))')
TR = 15
dsh.cell(TR, 1, 'Total')
for ci in range(2, 8): dsh.cell(TR, ci, f'=SUM({L(ci)}3:{L(ci)}14)')
dsh.cell(TR, 8, '=IF(COUNT($G$3:$G$14)=0,"",'
                'IF(ABS(SUMPRODUCT(($G$3:$G$14<>"")*$F$3:$F$14)-SUM($G$3:$G$14))<0.99,"Reconciled",'
                'SUMPRODUCT(($G$3:$G$14<>"")*$F$3:$F$14)-SUM($G$3:$G$14)))')
for r in range(3, TR + 1):
    for ci in range(1, 9):
        c = dsh.cell(r, ci)
        c.font = Font(F, 10, bold=(r == TR)); c.border = box
        c.alignment = Alignment(horizontal='center' if ci == 1 else 'right', vertical='center')
        if 2 <= ci <= 7: c.number_format = MONEY
        if ci == 8: c.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 7 and r < TR:
            c.fill = PatternFill('solid', fgColor=YELLOW); c.font = Font(F, 10, color='0000FF')
        elif r == TR: c.fill = PatternFill('solid', fgColor=GREY)
dsh['A17'] = ('P&L Total is the only figure you type on this sheet (shaded yellow). '
              'Variance reads "Reconciled" when the gap is under $0.99.')
dsh['A18'] = ('The Total row compares like with like: it measures only the months where a P&L Total '
              'has actually been entered, so a part-year does not show a false gap.')
for c in ('A17', 'A18'): dsh[c].font = Font(F, 9, italic=True, color='555555')

title(dsh, 'A21', 'Reconciliation — is every record reaching the monthly view?')
hdrow(dsh, 22, ['Source', 'Total in tracker', 'In FY27 months', 'No date',
                'Dated outside FY27', 'Adds up?'])
for i, s in enumerate(['Support', 'Production', 'Consulting', 'WIP']):
    r = 23 + i
    dsh.cell(r, 1, s)
    dsh.cell(r, 2, f'=SUMIF(Data!$B$2:$B$20000,$A{r},Data!$L$2:$L$200000)')
    dsh.cell(r, 3, f'=SUMIFS(Data!$L$2:$L$200000,Data!$B$2:$B$20000,$A{r},Data!$O$2:$O$20000,"In FY27")')
    dsh.cell(r, 4, f'=SUMIFS(Data!$L$2:$L$200000,Data!$B$2:$B$20000,$A{r},Data!$O$2:$O$20000,"No date")')
    dsh.cell(r, 5, f'=SUMIFS(Data!$L$2:$L$200000,Data!$B$2:$B$20000,$A{r},Data!$O$2:$O$20000,"Outside FY27")')
    dsh.cell(r, 6, f'=IF(ABS($B{r}-SUM($C{r}:$E{r}))<0.01,"OK","Check")')
dsh.cell(27, 1, 'Total')
for ci in range(2, 6): dsh.cell(27, ci, f'=SUM({L(ci)}23:{L(ci)}26)')
dsh.cell(27, 6, '=IF(ABS($C$27-$F$15)<0.01,"Matches monthly grid","Does not match monthly grid")')
for r in range(23, 28):
    for ci in range(1, 7):
        c = dsh.cell(r, ci)
        c.font = Font(F, 10, bold=(r == 27)); c.border = box
        c.alignment = Alignment(horizontal='center' if ci in (1, 6) else 'right', vertical='center')
        if 2 <= ci <= 5: c.number_format = MONEY
        if r == 27: c.fill = PatternFill('solid', fgColor=GREY)
dsh['A29'] = ('"No date" and "Dated outside FY27" are different problems. A blank date needs filling in. '
              'A date outside FY27 is a real date from another financial year — the record is valid, '
              'it just does not belong in this tracker\'s twelve months.')
dsh['A29'].font = Font(F, 9, italic=True, color='555555')
for col, w in zip('ABCDEFGH', [14, 17, 17, 17, 17, 17, 15, 24]): dsh.column_dimensions[col].width = w
dsh.freeze_panes = 'A3'

# ==========================================================  CHECKS
ck = wb.create_sheet('Checks')
ck.sheet_view.showGridLines = False
title(ck, 'A1', 'Data quality checks')
ck['A2'] = 'These are live formulas. They update as soon as you change anything on the Data tab.'
ck['A2'].font = Font(F, 9, italic=True, color='555555')
hdrow(ck, 4, ['#', 'What is being checked', 'Records', 'Value', 'Status', 'What to do'])
CHECKS = [
 ('Records with no date',
  '=COUNTIF(Data!$O$2:$O$20000,"No date")', '=SUMIF(Data!$O$2:$O$20000,"No date",Data!$L$2:$L$200000)',
  'Filter Date Check on the Data tab to "No date" and fill the dates in. Until you do, this money appears in no month.'),
 ('Records dated outside FY 2026-27',
  '=COUNTIF(Data!$O$2:$O$20000,"Outside FY27")', '=SUMIF(Data!$O$2:$O$20000,"Outside FY27",Data!$L$2:$L$200000)',
  'Real dates from another financial year. Decide whether they belong in this tracker at all, or move them to the FY26 file.'),
 ('WIP jobs that have also been invoiced',
  '=COUNTIF(Data!$P$2:$P$20000,"Job already invoiced")', '=SUMIF(Data!$P$2:$P$20000,"Job already invoiced",Data!$L$2:$L$200000)',
  'Filter Duplicate Check on the Data tab. Each of these needs a reversing WIP entry, or it is double counted.'),
 ('Departments not on the master list',
  '=SUMPRODUCT((Data!$D$2:$D$20000<>"")*ISNA(MATCH(Data!$D$2:$D$20000,Lists!$F$4:$F$10,0)))', '',
  'Either correct the record or add the department to the Lists tab.'),
 ('Records with money but no client',
  '=SUMPRODUCT((Data!$L$2:$L$20000<>0)*(Data!$L$2:$L$20000<>"")*(Data!$F$2:$F$20000=""))', '',
  'A revenue line with no client cannot be traced back to anything.'),
 ('Records with money but no job number',
  '=SUMPRODUCT((Data!$L$2:$L$20000<>0)*(Data!$L$2:$L$20000<>"")*(Data!$G$2:$G$20000=""))', '',
  'Without a job number the WIP duplicate check cannot see this record.'),
 ('Monthly grid vs total in tracker',
  '', '=Dashboard!$F$15-Dashboard!$C$27',
  'Should be zero. Anything else means the Dashboard and the Data tab disagree.'),
]
for i, (what, cnt, val, todo) in enumerate(CHECKS):
    r = 5 + i
    ck.cell(r, 1, i + 1); ck.cell(r, 2, what)
    if cnt: ck.cell(r, 3, cnt)
    if val: ck.cell(r, 4, val).number_format = MONEY
    ck.cell(r, 5, f'=IF(N($C{r})+ROUND(N($D{r}),2)=0,"OK","Review")')
    ck.cell(r, 6, todo)
    for ci in range(1, 7):
        c = ck.cell(r, ci); c.font = Font(F, 10); c.border = box
        c.alignment = Alignment(horizontal=('center' if ci in (1, 3, 5) else
                                            'right' if ci == 4 else 'left'),
                                vertical='center', wrap_text=(ci == 6))
    ck.row_dimensions[r].height = 30
for col, w in zip('ABCDEF', [5, 42, 10, 16, 11, 72]): ck.column_dimensions[col].width = w

dupe_jobs = {}
for rec in recs:
    if rec['source'] != 'WIP' and rec.get('job'):
        dupe_jobs.setdefault(str(rec['job']), []).append(rec)
rows = []
for rec in recs:
    if rec['source'] == 'WIP' and rec.get('job') and str(rec['job']) in dupe_jobs:
        for other in dupe_jobs[str(rec['job'])]:
            rows.append((str(rec['job']), rec.get('client'), rec.get('wip'),
                         other['source'], other.get('invval')))
R0 = 5 + len(CHECKS) + 2
title(ck, f'A{R0}', 'WIP jobs that already appear in an invoiced record')
ck.cell(R0 + 1, 1, f'Snapshot taken when this workbook was built. The live version is the '
                   f'Duplicate Check column on the Data tab.').font = Font(F, 9, italic=True, color='555555')
hdrow(ck, R0 + 2, ['Job Number', 'Client', 'WIP Movement', 'Also in', 'Invoice Value'])
for i, row in enumerate(rows):
    r = R0 + 3 + i
    for ci, x in enumerate(row, 1):
        c = ck.cell(r, ci, x); c.font = Font(F, 10); c.border = box
        c.alignment = Alignment(horizontal='center' if ci in (1, 4) else 'left' if ci == 2 else 'right')
        if ci in (3, 5): c.number_format = MONEY

# ==========================================================  LEGACY NOTES
ln = wb.create_sheet('Legacy Notes')
ln.sheet_view.showGridLines = False
title(ln, 'A1', 'Notes and comments recovered from the old workbook')
for i, t in enumerate([
    'In the old file these were Excel comments stuck to cell positions, not to jobs. Because the macro swapped the',
    'data underneath them every time you changed department, they had drifted away from the records they describe.',
    'They are listed here with their original cell reference so you can reattach each one to the right record by',
    'pasting it into the Notes column on the Data tab. Nothing has been guessed — no note has been assigned to a record.']):
    ln.cell(2 + i, 1, t).font = Font(F, 9, italic=True, color='555555')
hdrow(ln, 7, ['Original cell', 'Type', 'Author', 'Date', 'Note text'])
cm = sorted(D['comments'], key=lambda c: (int(''.join(ch for ch in c['ref'] if ch.isdigit()) or 0),
                                          ''.join(ch for ch in c['ref'] if ch.isalpha())))
w = 0
for i, c in enumerate(cm):
    if not c['text']: continue
    r = 8 + w; w += 1
    for ci, x in enumerate([c['ref'], c['kind'], c['who'], c['when'], c['text']], 1):
        cell = ln.cell(r, ci, x); cell.font = Font(F, 10); cell.border = box
        cell.alignment = Alignment(horizontal='center' if ci in (1, 2, 4) else 'left',
                                   vertical='top', wrap_text=(ci == 5))
    ln.row_dimensions[r].height = 14 * max(1, c['text'].count('\n') + 1)
for col, w2 in zip('ABCDE', [13, 11, 22, 12, 96]): ln.column_dimensions[col].width = w2
ln.freeze_panes = 'A8'

# ==========================================================  READ ME
rm = wb.create_sheet('Read Me', 0)
rm.sheet_view.showGridLines = False
rm.column_dimensions['A'].width = 3
rm.column_dimensions['B'].width = 112
def para(row, text, style='body'):
    c = rm.cell(row, 2, text)
    if style == 'h1':   c.font = Font(F, 16, bold=True, color=NAVY)
    elif style == 'h2': c.font = Font(F, 11, bold=True, color=NAVY)
    elif style == 'note': c.font = Font(F, 9, italic=True, color='555555')
    else: c.font = Font(F, 10)
    c.alignment = Alignment(vertical='top', wrap_text=True)
    rm.row_dimensions[row].height = 15 if style in ('h2',) else (26 if style == 'h1' else 13)
    return row + 1

CONTENT = [
 ('h1',  'FY 2026-27 Revenue Tracker'),
 ('note','Rebuilt from FY27_Revenue_Tracker_v2.xlsm. Your original file has not been changed.'),
 ('body',''),
 ('h2',  'What changed, and why'),
 ('body','The old workbook kept four separate hidden copies of your data off to the right of the sheet, and a macro'),
 ('body','wiped and re-typed the whole visible table every time you changed department. That is what made it slow, and'),
 ('body','it is why the Department box could disagree with the data actually on screen.'),
 ('body',''),
 ('body','This version has one table with every record in it, on the Data tab. There are no hidden copies and no macros.'),
 ('body','To see one department, click the filter arrow on the Department or Source column. It is instant, it cannot get'),
 ('body','out of step with what you are looking at, and you can filter on more than one thing at once.'),
 ('body',''),
 ('h2',  'The tabs'),
 ('body','Data            Every record, one row each. 181 rows brought across from the old file.'),
 ('body','Dashboard       Revenue by month and department, plus the reconciliation check. Type the P&L Total; nothing else.'),
 ('body','Checks          Live data quality checks. Look here first each month.'),
 ('body','Legacy Notes    The 156 comments recovered from the old file, for you to reattach.'),
 ('body','Lists           The dropdown values and the financial year dates. Change them here, not in the table.'),
 ('body',''),
 ('h2',  'How the Data tab works'),
 ('body','Source says which of the four old sheets a record came from, and drives the Dashboard columns.'),
 ('body','Department is the real department. Support records are all "Support"; WIP records keep the department of the'),
 ('body','underlying job. These are now one shared list on the Lists tab rather than four different lists.'),
 ('body',''),
 ('body','Grey columns are calculated — do not type in them:'),
 ('body','    Net Revenue      What rolls up to the Dashboard. WIP uses WIP Movement, Production uses Invoice Value less'),
 ('body','                     Discounts Included, Support and Consulting use Invoice Value. Same as the old file.'),
 ('body','    Margin           Production: Invoice Value less Cross Hire and Internal Labour. Consulting: Invoice Value'),
 ('body','                     less the three expense columns. Same as the old file.'),
 ('body','    Date Check       Flags "No date" or "Outside FY27" so nothing goes missing silently.'),
 ('body','    Duplicate Check  Flags a WIP job that has already been invoiced somewhere else.'),
 ('body',''),
 ('body','The Production and Consulting detail columns are grouped. Use the small + and - buttons above the column'),
 ('body','letters to collapse them when you do not need them.'),
 ('body',''),
 ('body','The table grows on its own. Type on the row underneath and it extends, formulas and all. There is no limit'),
 ('body','of 200 any more — the old file silently threw away anything past row 211. The Dashboard and Checks read the'),
 ('body','first 20,000 rows, roughly eighteen years at your current rate.'),
 ('body',''),
 ('h2',  'Things carried across that still need a decision'),
 ('body','None of these have been changed. They are flagged so you can decide.'),
 ('body',''),
 ('body','1.  26 records have no date. They cannot appear in any month until one is entered. See Checks.'),
 ('body','2.  16 Consulting records carry FY26 dates, from May 2025 to June 2026. In the old file these were lumped in'),
 ('body','    with the undated ones and labelled "Undated", which hid the problem. They are now shown separately.'),
 ('body','3.  Five WIP jobs have also been invoiced. One is a clean reversal, the rest are not. See Checks.'),
 ('body','4.  One record has the department "Onsite", which was not on any list in the old file. It has been kept and'),
 ('body','    added to the Lists tab so nothing was lost — reclassify it if it should be something else.'),
 ('body','5.  Margin on Production is calculated from Invoice Value, not Net Total, so discounts are not deducted before'),
 ('body','    margin. That is how the old file did it. Worth checking it is what you want.'),
 ('body',''),
 ('h2',  'One thing to know about the numbers'),
 ('body','The old file was showing stale Production figures — about $26,792 higher than its own formulas produced. Every'),
 ('body','figure here is calculated fresh, so July and August Production will read lower than you are used to. The lower'),
 ('body','numbers are the correct ones.'),
]
row = 2
for style, text in CONTENT: row = para(row, text, style)
rm['B2'].alignment = Alignment(vertical='center')

# ==========================================================  SAVE
for ws in wb.worksheets:
    ws.sheet_properties.tabColor = NAVY
OUT = 'FY27_Revenue_Tracker_v3.xlsx'
wb.save(OUT)
print('saved', OUT, '| data rows:', N, '| notes:', w)
