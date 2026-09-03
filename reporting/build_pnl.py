"""Build the CTS P&L reporting workbook.

One rule drives the design: the only sheets anyone types into are Actuals and
Budget. Every other sheet is formulas pointed at those two, so appending a
month updates the whole pack at once.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/home/user/Claude/reporting")
ACCOUNTS = json.loads((ROOT / "data/accounts.json").read_text())
VALID = json.loads((ROOT / "data/validation.json").read_text())

NAVY, MUTED = "1F4E79", "595959"
HEAD = PatternFill("solid", fgColor=NAVY)
BAND = PatternFill("solid", fgColor="D9E2F3")
INPUT = PatternFill("solid", fgColor="FFF2CC")
TOTAL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ARIAL = "Arial"
MONEY = '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'

GROUPS = ["Income", "Cost of Sales", "Other Income", "Expenses"]
DIVISIONS = ["Onsite", "Production", "Video", "Consulting", "Integration", "Admin", "Unallocated"]

# Australian financial year: July to June.
MONTHS = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]
QUARTERS = [("Q1 Jul-Sep", 0, 3), ("Q2 Oct-Dec", 3, 6), ("Q3 Jan-Mar", 6, 9), ("Q4 Apr-Jun", 9, 12)]
FY_LAST, FY_THIS = "FY26", "FY27"
LAST_YEARS = ["25", "25", "25", "25", "25", "25", "26", "26", "26", "26", "26", "26"]
THIS_YEARS = ["26", "26", "26", "26", "26", "26", "27", "27", "27", "27", "27", "27"]

DATA_FIRST_COL = 4                      # Actuals/Budget: months start at column D
LAST_COLS = [get_column_letter(DATA_FIRST_COL + i) for i in range(12)]        # D..O  FY26
THIS_COLS = [get_column_letter(DATA_FIRST_COL + 12 + i) for i in range(12)]   # P..AA FY27
DATA_FIRST_ROW = 4                      # Actuals/Budget: first account row


def title(ws, text, sub=None, width=14):
    ws["A1"] = text
    ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=ARIAL, size=9, color=MUTED)
    ws.sheet_view.showGridLines = False


def header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD
        c.border = BOX
        c.alignment = Alignment(horizontal="center" if i > 1 else "left", wrap_text=True)
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------- data sheets
def build_data_sheet(wb, name, note):
    ws = wb.create_sheet(name)
    title(ws, f"{name}: paste your Xero export here", note)
    labels = ["Account", "Group", "Division"]
    labels += [f"{m}-{y}" for m, y in zip(MONTHS, LAST_YEARS)]
    labels += [f"{m}-{y}" for m, y in zip(MONTHS, THIS_YEARS)]
    header_row(ws, 3, labels, widths=[34, 15, 13] + [12] * 24)
    for i, a in enumerate(ACCOUNTS):
        r = DATA_FIRST_ROW + i
        ws.cell(row=r, column=1, value=a["account"]).font = Font(name=ARIAL, size=9)
        ws.cell(row=r, column=2, value=a["group"]).font = Font(name=ARIAL, size=9, color=MUTED)
        ws.cell(row=r, column=3, value=a["division"]).font = Font(name=ARIAL, size=9, color=MUTED)
        for j in range(24):
            c = ws.cell(row=r, column=DATA_FIRST_COL + j)
            c.number_format = MONEY
            c.fill = INPUT
            c.font = Font(name=ARIAL, size=9)
            c.border = BOX
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(3 + 24)}{DATA_FIRST_ROW + len(ACCOUNTS) - 1}"
    return ws


# ------------------------------------------------------------- P&L statements
def build_pnl_sheet(wb, name, cols, fy, note):
    """A full P&L: accounts down, twelve months across, year total at the end."""
    ws = wb.create_sheet(name)
    title(ws, f"{name}: {fy} Profit & Loss", note)
    years = LAST_YEARS if fy == FY_LAST else THIS_YEARS
    header_row(ws, 3, ["Account"] + [f"{m}-{y}" for m, y in zip(MONTHS, years)] + [f"{fy} Total"],
               widths=[36] + [13] * 12 + [15])

    row = 4
    marks = {}

    def band(label):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=ARIAL, size=10, bold=True, color=NAVY)
        for j in range(1, 15):
            ws.cell(row=row, column=j).fill = BAND
        row += 1

    def accounts_of(group):
        nonlocal row
        start = row
        for a in [x for x in ACCOUNTS if x["group"] == group]:
            ws.cell(row=row, column=1, value=a["account"]).font = Font(name=ARIAL, size=9)
            for j, src in enumerate(cols):
                cell = ws.cell(row=row, column=2 + j)
                cell.value = f"=SUMIFS({data}!${src}:${src},{data}!$A:$A,$A{row})"
                cell.number_format = MONEY
                cell.font = Font(name=ARIAL, size=9)
            t = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
            t.number_format = MONEY
            t.font = Font(name=ARIAL, size=9, bold=True)
            row += 1
        return start, row - 1

    def total(label, formula_for):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=ARIAL, size=10, bold=True)
        for j in range(12):
            col = get_column_letter(2 + j)
            cell = ws.cell(row=row, column=2 + j, value=formula_for(col))
            cell.number_format = MONEY
            cell.font = Font(name=ARIAL, size=10, bold=True)
            cell.fill = TOTAL
        t = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        t.number_format = MONEY
        t.font = Font(name=ARIAL, size=10, bold=True)
        t.fill = TOTAL
        here = row
        row += 1
        return here

    data = "Actuals" if name != "Budget P&L" else "Budget"

    band("INCOME")
    inc = accounts_of("Income")
    marks["income"] = total("Total Income", lambda c: f"=SUM({c}{inc[0]}:{c}{inc[1]})")
    row += 1

    band("COST OF SALES")
    cos = accounts_of("Cost of Sales")
    marks["cos"] = total("Total Cost of Sales", lambda c: f"=SUM({c}{cos[0]}:{c}{cos[1]})")
    row += 1

    marks["gp"] = total("GROSS PROFIT", lambda c: f"={c}{marks['income']}-{c}{marks['cos']}")
    marks["gpm"] = total("Gross Profit %",
                         lambda c: f"=IFERROR({c}{marks['gp']}/{c}{marks['income']},0)")
    for j in list(range(2, 14)) + [14]:
        ws.cell(row=marks["gpm"], column=j).number_format = PCT
    row += 1

    band("OTHER INCOME")
    oth = accounts_of("Other Income")
    marks["other"] = total("Total Other Income", lambda c: f"=SUM({c}{oth[0]}:{c}{oth[1]})")
    row += 1

    band("EXPENSES")
    exp = accounts_of("Expenses")
    marks["exp"] = total("Total Expenses", lambda c: f"=SUM({c}{exp[0]}:{c}{exp[1]})")
    row += 1

    marks["np"] = total("NET PROFIT",
                        lambda c: f"={c}{marks['gp']}+{c}{marks['other']}-{c}{marks['exp']}")
    marks["npm"] = total("Net Profit %",
                         lambda c: f"=IFERROR({c}{marks['np']}/{c}{marks['income']},0)")
    for j in list(range(2, 14)) + [14]:
        ws.cell(row=marks["npm"], column=j).number_format = PCT

    ws.freeze_panes = "B4"
    return ws, marks


wb = Workbook()
wb.remove(wb.active)

actuals = build_data_sheet(
    wb, "Actuals",
    "Yellow cells only. Paste each month's figures from your Xero P&L export into the matching column. "
    "Nothing else on this sheet should be edited, and no other sheet should be typed into at all.")
budget = build_data_sheet(
    wb, "Budget",
    "Yellow cells only. Enter the budget for each account by month, in the same layout as Actuals.")

this_ws, this_marks = build_pnl_sheet(
    wb, "This Year", THIS_COLS, FY_THIS, "Calculated from the Actuals sheet. Do not type here.")
last_ws, last_marks = build_pnl_sheet(
    wb, "Last Year", LAST_COLS, FY_LAST, "Calculated from the Actuals sheet. Do not type here.")

wb.save(ROOT / "CTS P&L Reporting.xlsx")
print(f"built with {len(ACCOUNTS)} accounts")
print("This Year marker rows:", this_marks)
print("Last Year marker rows:", last_marks)


# --------------------------------------------------------- comparison sheets
def build_comparison(wb, name, left_sheet, right_sheet, left_label, right_label, note):
    """Row-for-row comparison. Both P&L sheets share a layout, so rows align exactly."""
    ws = wb.create_sheet(name)
    title(ws, name, note)
    header_row(ws, 3, ["Account", left_label, right_label, "Variance $", "Variance %"],
               widths=[38, 16, 16, 16, 12])
    src = wb[left_sheet]
    row = 4
    for r in range(4, src.max_row + 1):
        label = src.cell(row=r, column=1).value
        if label is None:
            row += 1
            continue
        is_pct = "%" in str(label)
        is_bold = str(label).isupper() or str(label).startswith("Total") or is_pct
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=ARIAL, size=10 if is_bold else 9, bold=is_bold,
                      color=NAVY if str(label).isupper() else "000000")
        for col, sheet in ((2, left_sheet), (3, right_sheet)):
            cell = ws.cell(row=row, column=col, value=f"='{sheet}'!N{r}")
            cell.number_format = PCT if is_pct else MONEY
            cell.font = Font(name=ARIAL, size=10 if is_bold else 9, bold=is_bold)
        v = ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        v.number_format = PCT if is_pct else MONEY
        v.font = Font(name=ARIAL, size=10 if is_bold else 9, bold=is_bold)
        p = ws.cell(row=row, column=5)
        p.value = "" if is_pct else f"=IFERROR(D{row}/ABS(C{row}),0)"
        p.number_format = PCT
        p.font = Font(name=ARIAL, size=10 if is_bold else 9, bold=is_bold)
        if is_bold:
            for j in range(1, 6):
                ws.cell(row=row, column=j).fill = TOTAL
        row += 1
    ws.freeze_panes = "B4"
    return ws


# ------------------------------------------------------------- division view
def build_divisions(wb):
    ws = wb.create_sheet("By Division")
    title(ws, "By Division: FY27",
          "Division is taken from each account's suffix (ONS, PRD, VID, CONS, INTEGRATION, ADMIN). "
          "Unallocated covers shared overheads that carry no division in the account name.")
    header_row(ws, 3, ["Division"] + [f"{m}-{y}" for m, y in zip(MONTHS, THIS_YEARS)] + ["FY27 Total"],
               widths=[22] + [13] * 12 + [15])

    row = 4
    blocks = {}
    for label, group in (("INCOME", "Income"), ("COST OF SALES", "Cost of Sales")):
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=ARIAL, size=10, bold=True, color=NAVY)
        for j in range(1, 15):
            ws.cell(row=row, column=j).fill = BAND
        row += 1
        blocks[group] = {}
        for d in DIVISIONS:
            ws.cell(row=row, column=1, value=d).font = Font(name=ARIAL, size=9)
            for j, src in enumerate(THIS_COLS):
                cell = ws.cell(row=row, column=2 + j)
                cell.value = (f'=SUMIFS(Actuals!${src}:${src},Actuals!$B:$B,"{group}",'
                              f'Actuals!$C:$C,$A{row})')
                cell.number_format = MONEY
                cell.font = Font(name=ARIAL, size=9)
            t = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
            t.number_format = MONEY
            t.font = Font(name=ARIAL, size=9, bold=True)
            blocks[group][d] = row
            row += 1
        row += 1

    c = ws.cell(row=row, column=1, value="GROSS PROFIT")
    c.font = Font(name=ARIAL, size=10, bold=True, color=NAVY)
    for j in range(1, 15):
        ws.cell(row=row, column=j).fill = BAND
    row += 1
    for d in DIVISIONS:
        ws.cell(row=row, column=1, value=d).font = Font(name=ARIAL, size=9)
        inc, cos = blocks["Income"][d], blocks["Cost of Sales"][d]
        for j in range(12):
            col = get_column_letter(2 + j)
            cell = ws.cell(row=row, column=2 + j, value=f"={col}{inc}-{col}{cos}")
            cell.number_format = MONEY
            cell.font = Font(name=ARIAL, size=9)
        t = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        t.number_format = MONEY
        t.font = Font(name=ARIAL, size=9, bold=True)
        row += 1
    ws.freeze_panes = "B4"
    return ws


# -------------------------------------------------------------------- summary
def build_summary(wb, this_marks, last_marks):
    ws = wb.create_sheet("Summary", 0)
    title(ws, "Summary: FY27",
          "Everything here is calculated. The only sheets you type into are Actuals and Budget.")

    lines = [("Total Income", "income", MONEY), ("Total Cost of Sales", "cos", MONEY),
             ("Gross Profit", "gp", MONEY), ("Gross Profit %", "gpm", PCT),
             ("Total Other Income", "other", MONEY), ("Total Expenses", "exp", MONEY),
             ("Net Profit", "np", MONEY), ("Net Profit %", "npm", PCT)]

    ws["A4"] = "MONTHLY"
    ws["A4"].font = Font(name=ARIAL, size=11, bold=True, color=NAVY)
    header_row(ws, 5, ["  "] + [f"{m}-{y}" for m, y in zip(MONTHS, THIS_YEARS)] + ["FY27 Total"],
               widths=[26] + [13] * 12 + [15])
    row = 6
    for label, key, fmt in lines:
        ws.cell(row=row, column=1, value=label).font = Font(name=ARIAL, size=10, bold=True)
        for j in range(12):
            col = get_column_letter(2 + j)
            cell = ws.cell(row=row, column=2 + j, value=f"='This Year'!{col}{this_marks[key]}")
            cell.number_format = fmt
            cell.font = Font(name=ARIAL, size=10)
        t = ws.cell(row=row, column=14, value=f"='This Year'!N{this_marks[key]}")
        t.number_format = fmt
        t.font = Font(name=ARIAL, size=10, bold=True)
        t.fill = TOTAL
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="QUARTERLY").font = Font(name=ARIAL, size=11, bold=True, color=NAVY)
    row += 1
    header_row(ws, row, ["  "] + [q for q, _, _ in QUARTERS] + ["FY27 Total"])
    qhead = row
    row += 1
    for i, (label, key, fmt) in enumerate(lines):
        monthly_row = 6 + i
        ws.cell(row=row, column=1, value=label).font = Font(name=ARIAL, size=10, bold=True)
        for qi, (_, a, b) in enumerate(QUARTERS):
            cell = ws.cell(row=row, column=2 + qi)
            if fmt is PCT:
                # A margin is not the sum of three margins; rebuild it from the quarter.
                inc_row = 6
                num_row = 6 + (2 if key == "gpm" else 6)
                cell.value = (f"=IFERROR(SUM({get_column_letter(2+a)}{num_row}:"
                              f"{get_column_letter(1+b)}{num_row})/SUM("
                              f"{get_column_letter(2+a)}{inc_row}:{get_column_letter(1+b)}{inc_row}),0)")
            else:
                cell.value = (f"=SUM({get_column_letter(2+a)}{monthly_row}:"
                              f"{get_column_letter(1+b)}{monthly_row})")
            cell.number_format = fmt
            cell.font = Font(name=ARIAL, size=10)
        t = ws.cell(row=row, column=14, value=f"='This Year'!N{this_marks[key]}")
        t.number_format = fmt
        t.font = Font(name=ARIAL, size=10, bold=True)
        t.fill = TOTAL
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="YEAR ON YEAR").font = Font(name=ARIAL, size=11, bold=True, color=NAVY)
    row += 1
    header_row(ws, row, ["  ", "FY27", "FY26", "Variance $", "Variance %"])
    row += 1
    for label, key, fmt in lines:
        ws.cell(row=row, column=1, value=label).font = Font(name=ARIAL, size=10, bold=True)
        a = ws.cell(row=row, column=2, value=f"='This Year'!N{this_marks[key]}")
        b = ws.cell(row=row, column=3, value=f"='Last Year'!N{last_marks[key]}")
        d = ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        p = ws.cell(row=row, column=5)
        p.value = "" if fmt is PCT else f"=IFERROR(D{row}/ABS(C{row}),0)"
        for cell, f in ((a, fmt), (b, fmt), (d, fmt), (p, PCT)):
            cell.number_format = f
            cell.font = Font(name=ARIAL, size=10)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="RECONCILIATION CHECK").font = Font(name=ARIAL, size=11, bold=True, color=NAVY)
    row += 1
    ws.cell(row=row, column=1, value=(
        "FY26 figures pulled directly from Xero on 3 September 2026. Once you have pasted FY26 into "
        "Actuals, the difference column should read zero. Anything else means the paste is incomplete "
        "or misaligned.")).font = Font(name=ARIAL, size=9, italic=True, color=MUTED)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 30
    row += 1
    header_row(ws, row, ["  ", "Workbook FY26", "Xero FY26", "Difference", ""])
    row += 1
    fy = VALID["periods"]["FY26"]
    for label, key, xero in (("Total Income", "income", fy["total_income"]),
                             ("Total Cost of Sales", "cos", fy["total_cost_of_sales"]),
                             ("Gross Profit", "gp", fy["gross_profit"]),
                             ("Total Other Income", "other", fy["total_other_income"]),
                             ("Total Expenses", "exp", fy["total_expenses"]),
                             ("Net Profit", "np", fy["net_profit"])):
        ws.cell(row=row, column=1, value=label).font = Font(name=ARIAL, size=10, bold=True)
        a = ws.cell(row=row, column=2, value=f"='Last Year'!N{last_marks[key]}")
        b = ws.cell(row=row, column=3, value=xero)
        b.font = Font(name=ARIAL, size=10, color="0000FF")   # hardcoded input, per convention
        d = ws.cell(row=row, column=4, value=f"=ROUND(B{row}-C{row},2)")
        for cell in (a, b, d):
            cell.number_format = MONEY
        a.font = Font(name=ARIAL, size=10)
        d.font = Font(name=ARIAL, size=10, bold=True)
        row += 1
    return ws


build_comparison(wb, "Year on Year", "This Year", "Last Year", "FY27", "FY26",
                 "FY27 against FY26, account by account. Calculated from the two P&L sheets.")
build_divisions(wb)
build_summary(wb, this_marks, last_marks)
wb.save(ROOT / "CTS P&L Reporting.xlsx")
print("sheets:", wb.sheetnames)


budget_ws, budget_marks = build_pnl_sheet(
    wb, "Budget P&L", THIS_COLS, FY_THIS, "Calculated from the Budget sheet. Do not type here.")
build_comparison(wb, "Budget vs Actual", "This Year", "Budget P&L", "Actual FY27", "Budget FY27",
                 "Actual against budget, account by account. A positive variance means actual is "
                 "above budget, which is good for income and bad for costs.")


def build_readme(wb):
    ws = wb.create_sheet("Read Me", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    def line(row, text, size=10, bold=False, color="000000", height=None):
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if height:
            ws.row_dimensions[row].height = height

    line(2, "CTS P&L Reporting", 18, True, NAVY)
    line(3, "Corporate Technology Services Pty Ltd    Financial year 1 July to 30 June    AUD",
         9, False, MUTED)

    line(5, "The one rule", 12, True, NAVY)
    line(6, "You only ever type into two sheets: Actuals and Budget. Every other sheet is "
            "calculated and will update itself. If you find yourself typing a number anywhere "
            "else, something has gone wrong.", height=42)

    line(8, "Adding a month", 12, True, NAVY)
    line(9, "1.  In Xero, run the Profit and Loss report for the month and export it to Excel.", height=18)
    line(10, "2.  Open the Actuals sheet here and find the column for that month.", height=18)
    line(11, "3.  Paste the figures into the yellow cells in that column, matching account to account.", height=18)
    line(12, "4.  That is it. Every other sheet has already updated.", height=18)

    line(14, "Checking your paste worked", 12, True, NAVY)
    line(15, "The Summary sheet ends with a Reconciliation Check. It compares the workbook's FY26 "
             "totals against the figures pulled straight from Xero on 3 September 2026. Once FY26 "
             "is pasted in, the Difference column should read zero. Anything else means the paste "
             "is incomplete or a row has slipped out of alignment.", height=56)

    line(17, "What is on each sheet", 12, True, NAVY)
    rows = [
        ("Summary", "Monthly, quarterly and year on year on one page, plus the reconciliation check."),
        ("Actuals", "Your figures. 190 accounts down, 24 months across, FY26 and FY27. Paste here."),
        ("Budget", "The budget, in exactly the same layout. Enter it here."),
        ("This Year", "Full FY27 P&L, month by month."),
        ("Last Year", "Full FY26 P&L, month by month."),
        ("Budget P&L", "The budget in P&L form, so it lines up against the other two."),
        ("Year on Year", "FY27 against FY26, account by account, with variances."),
        ("Budget vs Actual", "Actual against budget, account by account, with variances."),
        ("By Division", "Income, cost of sales and gross profit split by division, month by month."),
    ]
    r = 18
    for name, desc in rows:
        c = ws.cell(row=r, column=2, value=f"{name}          {desc}")
        c.font = Font(name=ARIAL, size=10)
        r += 1

    line(r + 1, "Where the account list came from", 12, True, NAVY)
    line(r + 2, "All 190 accounts were read directly from your Xero chart of accounts on 3 September "
                "2026, so the names match Xero exactly and your export should paste straight in. "
                "Many carry no balance in a given month; that is normal and they are kept so the "
                "P&L stays complete.", height=56)

    line(r + 4, "Divisions", 12, True, NAVY)
    line(r + 5, "Division is derived from the account name suffix: ONS is Onsite, PRD is Production, "
                "VID is Video, CONS is Consulting, plus Integration and Admin. 88 accounts carry a "
                "division. The other 102 are shared overheads and show as Unallocated.", height=42)

    line(r + 7, "When FY28 starts", 12, True, NAVY)
    line(r + 8, "This workbook holds FY26 and FY27. When FY28 begins, send it back to me and I will "
                "roll it forward rather than you rebuilding it by hand.", height=30)
    return ws


build_readme(wb)
wb.save(ROOT / "CTS P&L Reporting.xlsx")
print("final sheets:", wb.sheetnames)
